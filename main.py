#!/usr/bin/env python

# Class Info: CISP253-23151 (Fall 2023)
# Author: Benjamin Gombas
# Contact: bgombas@email.davenport.edu
# Date: December 13, 2023
# Program name: mainmain.py

"""
This program reads in data from an Excel file and stores it in a SQLite db. The program then reads in data from the db
and displays it as a table using TKinter. The program then allows the user to filter the data.
"""
import os
import tkinter as tk
from tkinter import *
from tkinter import ttk
from database import Database
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))


##########
# The following area should only contain Functions and Classes
##########

class DbBrowser:
    """
    Overall class to manage the insurance viewer behavior

    Attributes:

    - data :    :class:`list` --> A list of data from from the sqlite database
    - database :    :class:`database.Database` --> An instance of the database class
    - entry_count :    :class:`tkinter.Label` --> The label that displays how many results are showing
    - entry_count_label :    :class:`tkinter.Label` --> The label that labels the entry count
    - entry_count_var :    :class:`tkinter.StringVar` --> The variable of the entry count, allowing for the label to be changed
    - field_mapping :    :class:`dict` --> A dictionary to easily take the table's column and get the sqlite field equivalent
    - filtered_data :    :class:`list` --> A list of data that has been filtered based on the selected filters
    - primary_filter :    :class:`tkinter.StringVar` --> String variable for the primary filter
    - primary_filter_label :    :class:`tkinter.Label` --> Label for the primary filter
    - primary_filter_menu :    :class:`tkinter.OptionMenu` --> The drop down menu allowing users to select a filter
    - root :    :class:`tk.Tk` --> Instance of tk
    - secondary_filter :    :class:`tkinter.StringVar` --> The string variable for the secondary filter
    - secondary_filter_label :    :class:`tkinter.Label` --> The label for the secondary filter
    - secondary_filter_menu :    :class:`tkinter.OptionMenu` --> The drop down menu for the secondary filter
    - secondary_menu_options :    :class:`list` --> A list containing the possible options for the secondary filter
    - table :    :class:`tkinter.ttk.Treeview` --> A treeview instance that is used to create a table
    - title_label :    :class:`tkinter.Label` --> A label that titles the screen

    Methods:
    - _import_excel() --> Import data from the excel file. convert y/n to 1/0
    - _update_db() --> Create an instance of Database, pull data from the excel file to ensure the most up-to-date data, store the data in the database
    - _window_setup() --> Set up basic parts of the TKinter window, such as the window title, the size, and resizeability
    - _display_tkinter_widgets() --> Display all of the tkinter widgets, including labels, the drop down menus, and the table
    - _filter_widgets() --> Set up the two filter widgets
    - _update_second_dropdown(*args) --> When a primary filter is selected, update the options of the secondary filter.
    - _update_data(*args) --> Update the table to show the filtered data
    - _create_table_rows(data) --> Generate rows for the treeview table
    - _data_table(data) --> Draw the data table
    """
    def __init__(self):
        # Put the current version of data.xlsx into the database
        self._update_db()
        # Read and store the data from the insurance table
        self.data = self.database.read_data()
        # Initialize things such as screen size, window title, and resizability
        self._window_setup()
        # Display all of the tkinter widgets
        self._display_tkinter_widgets()
        # Main tkinter loop
        self.root.mainloop()

    # This function has no use outside of being used by this class, so I am making it a static method to make the
    # squiggly lines go away telling me to make it a function instead of a method
    @staticmethod
    def _import_excel():
        """
        Import data from the excel file. convert y/n to 1/0
        :return sheet: The data from the excel file
        """
        excel_file = os.path.join(os.getcwd(), "data.xlsx")
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        # Iterate through rows in the Excel sheet
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.min_column,
                                   max_col=sheet.max_column):
            # Replace "Y" with 1 and "N" with 0 in the earthquake and flood table
            if row[8].value == "Y":
                row[8].value = 1
            elif row[8].value == "N":
                row[8].value = 0

            if row[9].value == "Y":
                row[9].value = 1
            elif row[9].value == "N":
                row[9].value = 0
        return sheet

    def _update_db(self):
        """
        Create an instance of Database, pull data from the excel file to ensure the most up-to-date data, store the data in the database
        :return None:
        """
        # Create and store an instance of the Database class
        self.database = Database()
        # Read in the Excel data
        fresh_data = self._import_excel()
        # Write the Excel data to the insurance table
        self.database.write_data(fresh_data)

    def _window_setup(self):
        """
        Set up basic parts of the TKinter window, such as the window title, the size, and resizeability
        :return None:
        """
        # Create an instance of TKinter
        self.root = tk.Tk()
        # Set the title of the window
        self.root.title("CISP253 - Final Project")
        # Set the size of the window
        self.root.minsize(width=1280, height=720)
        # Remove the ability to resize the window because I used absolute coords to place the scroll bar for the data
        self.root.resizable(width=0, height=0)

    def _display_tkinter_widgets(self):
        """
        Display all of the tkinter widgets, including labels, the drop down menus, and the table
        :return None:
        """
        root = self.root

        # Create a label for the window
        self.title_label = tk.Label(root, text="Insurance Data Viewer", font=("Helvetica", 42, "bold"))
        # Place the label at the top middle of the screen
        self.title_label.pack(side="top", pady=20)

        # Draw the dropdown menus
        self._filter_widgets()

        # Draw the number of entries labels
        self.entry_count_label = tk.Label(root, text="# of Entries", font=("Helvetica", 16))
        self.entry_count_label.place(x=1000, y=170)

        # Draw the number of entries
        self.entry_count_var = tk.StringVar()
        self.entry_count_var.set(str(len(self.data)))
        self.entry_count = tk.Label(root, textvariable=self.entry_count_var, font=("Helvetica", 12, "bold"))
        self.entry_count.place(x=1045, y=200)

        # Draw the data table
        self._data_table()

    def _filter_widgets(self):
        """
        Set up the two filter widgets
        :return None:
        """
        root = self.root

        # Dictionary mapping table column names to SQLite fields
        self.field_mapping = {
            "All": "*",
            "Policy": "policy",
            "Expiry": "expiry",
            "Location": "location",
            "State": "state",
            "Region": "region",
            "Insured Value": "insurance_value",
            "Construction": "construction",
            "Business Type": "business_type",
            "Earthquake": "earthquake",
            "Flood": "flood"
        }
        # Create and set up the primary dropdown menu
        self.primary_filter = StringVar(root)
        self.primary_filter.set("All")
        self.primary_filter.trace_add("write", self._update_second_dropdown)
        self.primary_filter_menu = OptionMenu(root, self.primary_filter, *self.field_mapping.keys())
        self.primary_filter_menu.place(x=180, y=200, width=165)

        # Create and set up the second dropdown menu
        self.secondary_filter = tk.StringVar()
        self.secondary_filter.trace_add("write", self._update_data)
        self.secondary_filter_menu = tk.OptionMenu(root, self.secondary_filter, "")
        self.secondary_filter_menu.place(x=550, y=200, width=195)

        # Create a label for the primary filter
        self.primary_filter_label = tk.Label(root, text="Select First Filter", font=("Helvetica", 16))
        # Place the label above the primary filter menu
        self.primary_filter_label.place(x=180, y=170)

        # Create a label for the secondary filter
        self.secondary_filter_label = tk.Label(root, text="Select Second Filter", font=("Helvetica", 16))
        # Place the label above the secondary filter menu
        self.secondary_filter_label.place(x=550, y=170)

    def _update_second_dropdown(self, *args):
        """
        When a primary filter is selected, update the options of the secondary filter.
        :param args:
        :return None:
        """
        # Get the currently selected primary filter
        selected_field = self.primary_filter.get()
        # Get the SQLite field associated with the selected display name
        sqlite_field = self.field_mapping[selected_field]
        # If the filter is "All", disable the secondary filter
        if selected_field == "All":
            self.secondary_filter.set("")
            self.secondary_filter_menu.config(state="disabled")
        else:
            # Change the binary options of earthquake and flood to the more user friendly yes/no
            if sqlite_field == "earthquake" or sqlite_field == "flood":
                self.secondary_menu_options = ["Yes", "No"]
            else:
                # Pull the unique values for the field from the database
                self.secondary_menu_options = self.database.read_unique_data(sqlite_field)

            # Enable the second dropdown menu and update its options
            self.secondary_filter_menu.config(state="normal")

            # Sort the values alphabetically
            self.secondary_menu_options.sort()

            # Update the options
            # Retrieve the menu object
            menu = self.secondary_filter_menu["menu"]
            # Clear the existing options in the menu
            menu.delete(0, "end")
            # Loop over the unique values
            for value in self.secondary_menu_options:
                #  Associate each menu option with the self.secondary_filter and set its value to the current value.
                #  This is a way to update the variable (self.secondary_filter) when an option is selected from the
                #  dropdown menu
                menu.add_command(label=value, command=tk._setit(self.secondary_filter, value))

    def _update_data(self, *args):
        """
        Update the table to show the filtered data
        :param args:
        :return None:
        """
        # if the secondary filter is "yes", replace it with 1 (the db value) and "no" with 0
        if self.secondary_filter.get() == "Yes":
            self.filtered_data = self.database.read_filtered_data(self.field_mapping[self.primary_filter.get()], 1)
        elif self.secondary_filter.get() == "No":
            self.filtered_data = self.database.read_filtered_data(self.field_mapping[self.primary_filter.get()], 0)
        # Otherwise, just pull the data from the db based on the selected filters
        else:
            self.filtered_data = self.database.read_filtered_data(self.field_mapping[self.primary_filter.get()],
                                                                  self.secondary_filter.get())
        # If the table exists
        if self.table:
            # Wipe all the rows
            self.table.delete(*self.table.get_children())
            # Generate the new table rows
            self._create_table_rows(self.filtered_data)
            # Set the entry count based on how many rows are found
            self.entry_count_var.set(str(len(self.filtered_data)))

    def _create_table_rows(self, data):
        """
        Generate rows for the treeview table
        :param data: a list of tuples
        :return None:
        """
        for row in data:
            # Remove the last 2 items in the row
            cut_tuple = row[:-2]
            # Replace 1s and 0s with checked or unchecked boxes
            if (row[9] == 1) and (row[10] == 1):
                cut_tuple += ("☑", "☑")
            elif row[9] == 0 and row[10] == 0:
                cut_tuple += ("☐", "☐")
            elif row[9] == 1 and row[10] == 0:
                cut_tuple += ("☑", "☐")
            elif row[9] == 0 and row[10] == 1:
                cut_tuple += ("☐", "☑")
            # Insert the row into the table
            self.table.insert("", "end", text=cut_tuple[0], values=cut_tuple[1:])

    def _data_table(self):
        """
        Draw the data table
        :return None:
        """
        # Check to see if the table exists already
        root = self.root
        # Calculate the number of rows that can fit in 2/5ths of the screen height
        height_of_table = ((root.winfo_reqheight() // 5) * 2) // 5

        # Create an instance of Treeview that will be used as a table
        self.table = ttk.Treeview(root, height=height_of_table)
        table = self.table

        # Make a scroll bar for the data
        scroll = ttk.Scrollbar(root, orient="vertical", command=table.yview)
        scroll.place(x=1265, y=399, height=720 - 399)
        table.configure(yscrollcommand=scroll.set)

        # Name the columns of the table
        table['columns'] = ("Policy", "Expiry", "Location", "State", "Region", "Insured Value", "Construction",
                            "Business Type", "Earthquake", "Flood")
        table.column("#0", width=110, anchor=CENTER)
        table.heading("#0", text="#")
        for col in table["columns"]:
            table.column(col, width=110, anchor=CENTER)
            table.heading(col, text=col.title())

        # Create the rows
        self._create_table_rows(self.data)

        # Pack the table to the bottom of the screen and let it span across the x axis and fill from it's height
        table.pack(fill="both", side="bottom")


#######################################
# All supporting functions and classes
# exist above this line
# ====================================
# Program Starts Here
def main():
    """
    Create an instance of DbBrowser, the class that runs the program.
    """
    DbBrowser()


# ===============================
# No extra Code beyond this point
# This code is required for the main() function to work
if __name__ == "__main__":
    main()
# EOF #
